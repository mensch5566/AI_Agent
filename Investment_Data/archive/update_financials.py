"""
update_financials.py
────────────────────────────────────────────────────────────────
Micron Technology (MU) 財務數據更新腳本
每次有新財報（10-Q 或 10-K）時執行此腳本，將新季度數據
合併進 MU_Q1FY2026_10Q.json，並重新產生 MU_Q1FY2026_10Q.xlsx。

使用方式：
    1. 在下方 NEW_PERIOD 區塊填入新季度的數據
    2. 執行：python3 update_financials.py
    3. 確認輸出無誤後，兩個檔案會自動更新

注意事項：
    - 不知道的數值填 None，腳本會顯示 "—" 並跳過 Long Format
    - periods 陣列由舊到新排列（腳本自動插入正確位置）
    - 每次執行前建議先備份現有檔案
────────────────────────────────────────────────────────────────
"""

import json, os, shutil
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR       = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(DIR, "MU_Q1FY2026_10Q.json")
XLSX_PATH = os.path.join(DIR, "MU_Q1FY2026_10Q.xlsx")

# ══════════════════════════════════════════════════════════════
# ▼▼▼  每次更新只需修改這個區塊  ▼▼▼
# ══════════════════════════════════════════════════════════════

NEW_PERIOD = "Q2_FY2026"   # 期間代號，格式：Q{1-4}_FY{年份}

NEW_FILING = {
    "form":             "10-Q",          # 10-Q 或 10-K
    "period_end":       "2026-02-26",    # 季末日
    "filing_date":      "2026-03-XX",    # 申報日（填後更新）
    "accession_number": "",              # SEC accession number
    "source_url":       "",              # SEC EDGAR URL
}

NEW_INCOME_STATEMENT = {
    "revenue":                            None,
    "cost_of_goods_sold":                 None,
    "gross_margin":                       None,
    "gross_margin_pct":                   None,  # 小數，如 0.56 表示 56%
    "research_and_development":           None,
    "selling_general_administrative":     None,
    "other_operating_income_expense_net": None,
    "operating_income":                   None,
    "operating_margin_pct":               None,
    "interest_income":                    None,
    "interest_expense":                   None,
    "interest_income_net":                None,
    "other_nonoperating_income_expense":  None,
    "income_before_taxes":                None,
    "income_tax_provision":               None,
    "effective_tax_rate":                 None,
    "equity_in_net_income_of_investees":  None,
    "net_income":                         None,
    "net_margin_pct":                     None,
    "eps_basic":                          None,
    "eps_diluted":                        None,
    "shares_basic_millions":              None,
    "shares_diluted_millions":            None,
}

NEW_BALANCE_SHEET = {
    "assets": {
        "cash_and_cash_equivalents":         None,
        "short_term_investments":            None,
        "receivables":                       None,
        "inventories":                       None,
        "other_current_assets":              None,
        "total_current_assets":              None,
        "long_term_marketable_investments":  None,
        "property_plant_equipment_net":      None,
        "operating_lease_right_of_use":      None,
        "intangible_assets":                 None,
        "deferred_tax_assets":               None,
        "goodwill":                          None,
        "other_noncurrent_assets":           None,
        "total_assets":                      None,
    },
    "liabilities": {
        "accounts_payable_and_accrued_expenses":  None,
        "current_debt":                           None,
        "other_current_liabilities":              None,
        "total_current_liabilities":              None,
        "long_term_debt":                         None,
        "noncurrent_operating_lease_liabilities": None,
        "noncurrent_unearned_gov_incentives":     None,
        "other_noncurrent_liabilities":           None,
        "total_liabilities":                      None,
    },
    "equity": {
        "common_stock":                               None,
        "additional_capital":                         None,
        "retained_earnings":                          None,
        "treasury_stock":                             None,
        "accumulated_other_comprehensive_income_loss":None,
        "total_equity":                               None,
        "total_liabilities_and_equity":               None,
    }
}

NEW_CASH_FLOW = {
    "operating_activities": {
        "net_income":                             None,
        "depreciation_and_amortization":          None,
        "stock_based_compensation":               None,
        "change_in_receivables":                  None,
        "change_in_inventories":                  None,
        "change_in_accounts_payable_accrued":     None,
        "change_in_other_current_liabilities":    None,
        "change_in_other_noncurrent_liabilities": None,
        "other":                                  None,
        "net_cash_from_operating_activities":     None,
    },
    "investing_activities": {
        "capital_expenditures":                   None,
        "purchases_of_afs_securities":            None,
        "proceeds_from_government_incentives":    None,
        "proceeds_from_maturities_sales_of_afs":  None,
        "other":                                  None,
        "net_cash_used_for_investing_activities": None,
    },
    "financing_activities": {
        "repayments_of_debt":                     None,
        "repurchases_stock_withholdings":         None,
        "repurchases_stock_program":              None,
        "dividends_paid":                         None,
        "other":                                  None,
        "net_cash_used_for_financing_activities": None,
    },
    "fx_effect_on_cash":  None,
    "net_change_in_cash": None,
    "beginning_cash":     None,
    "ending_cash":        None,
    "free_cash_flow":     None,  # = OCF + CapEx（CapEx 為負數）
}

NEW_REVENUE_BY_PRODUCT = {
    "DRAM":              None,
    "NAND":              None,
    "Other_NOR":         None,
    "total":             None,
    "DRAM_mix_pct":      None,
    "NAND_mix_pct":      None,
    "Other_NOR_mix_pct": None,
}

NEW_SEGMENT_DATA = {
    "CMBU": {
        "revenue":               None,
        "revenue_mix_pct":       None,
        "operating_income":      None,
        "operating_margin_pct":  None,
        "depreciation_amortization": None,
        "goodwill":              None,
    },
    "CDBU": {
        "revenue":               None,
        "revenue_mix_pct":       None,
        "operating_income":      None,
        "operating_margin_pct":  None,
        "depreciation_amortization": None,
        "goodwill":              None,
    },
    "MCBU": {
        "revenue":               None,
        "revenue_mix_pct":       None,
        "operating_income":      None,
        "operating_margin_pct":  None,
        "depreciation_amortization": None,
        "goodwill":              None,
    },
    "AEBU": {
        "revenue":               None,
        "revenue_mix_pct":       None,
        "operating_income":      None,
        "operating_margin_pct":  None,
        "depreciation_amortization": None,
        "goodwill":              None,
    },
    "total": {
        "revenue":          None,
        "operating_income": None,
        "unallocated":      None,
    }
}

# Notable Events：有就加，沒有就留空 list
NEW_NOTABLE_EVENTS = [
    # {
    #     "period":      NEW_PERIOD,
    #     "category":    "Debt",
    #     "date":        "2026-01-01",
    #     "description": "...",
    # },
]

# ▲▲▲  填完上面，執行腳本即可  ▲▲▲
# ══════════════════════════════════════════════════════════════


def merge_period(data, period, filing, is_dict, bs, cf, rb, seg_new, events):
    """將新 period 的數據合併進現有 data dict"""

    # 0. filing registry
    data["filings"][period] = filing

    # 1. periods 陣列（維持由舊到新）
    if period not in data["metadata"]["periods"]:
        data["metadata"]["periods"].append(period)
    data["metadata"]["last_updated"] = str(date.today())

    # 2. income statement
    for key, val in is_dict.items():
        if key not in data["income_statement"]:
            data["income_statement"][key] = {}
        data["income_statement"][key][period] = val

    # 3. balance sheet
    for sec in ["assets", "liabilities", "equity"]:
        for key, val in bs[sec].items():
            if key not in data["balance_sheet"][sec]:
                data["balance_sheet"][sec][key] = {}
            data["balance_sheet"][sec][key][period] = val

    # 4. cash flow
    for sec in ["operating_activities", "investing_activities", "financing_activities"]:
        for key, val in cf[sec].items():
            if key not in data["cash_flow_statement"][sec]:
                data["cash_flow_statement"][sec][key] = {}
            data["cash_flow_statement"][sec][key][period] = val
    for summary_key in ["fx_effect_on_cash","net_change_in_cash","beginning_cash","ending_cash","free_cash_flow"]:
        data["cash_flow_statement"][summary_key][period] = cf[summary_key]

    # 5. revenue by product
    for key, val in rb.items():
        if key not in data["revenue_by_product"]:
            data["revenue_by_product"][key] = {}
        data["revenue_by_product"][key][period] = val

    # 6. segment
    for sk in ["CMBU","CDBU","MCBU","AEBU"]:
        for metric, val in seg_new[sk].items():
            if metric not in data["segment_data"][sk]:
                data["segment_data"][sk][metric] = {}
            data["segment_data"][sk][metric][period] = val
    for metric, val in seg_new["total"].items():
        data["segment_data"]["total"][metric][period] = val

    # 7. notable events
    data["notable_events"].extend(events)

    return data


def rebuild_xlsx(data, xlsx_path):
    """根據最新 data 重建 XLSX（與原始腳本邏輯相同）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    PERIODS   = data["metadata"]["periods"]
    P_ENDS    = {p: data["filings"][p]["period_end"] for p in PERIODS}
    P_FORMS   = {p: data["filings"][p]["form"]       for p in PERIODS}
    N_PERIODS = len(PERIODS)
    NCOLS     = N_PERIODS + 2

    H_FILL=PatternFill("solid",fgColor="1F4E79"); SEC_FILL=PatternFill("solid",fgColor="D6E4F0")
    TOT_FILL=PatternFill("solid",fgColor="FFF2CC"); ALT_FILL=PatternFill("solid",fgColor="F2F8FD")
    W_FILL=PatternFill("solid",fgColor="FFFFFF");  NEW_FILL=PatternFill("solid",fgColor="E2EFDA")
    H_FONT=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
    T_FONT=Font(name="Calibri",bold=True,color="1F4E79",size=13)
    B_FONT=Font(name="Calibri",size=10); BD_FONT=Font(name="Calibri",bold=True,size=10)
    TOT_FONT=Font(name="Calibri",bold=True,color="7B3F00",size=10)
    NEW_FONT=Font(name="Calibri",bold=True,color="375623",size=9)
    thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
    CA=Alignment(horizontal="center",vertical="center",wrap_text=True)
    RA=Alignment(horizontal="right",vertical="center"); LA=Alignment(horizontal="left",vertical="center")

    def cs(c,value,font=B_FONT,fill=W_FILL,align=RA,num_fmt=None):
        c.value=value; c.font=font; c.fill=fill; c.alignment=align; c.border=BORDER
        if num_fmt and isinstance(value,(int,float)): c.number_format=num_fmt

    def phdr(p): return f"{p}\n{P_ENDS[p]}\n({P_FORMS[p]})"

    def write_col_hdrs(ws,row):
        cs(ws.cell(row=row,column=1),"",font=H_FONT,fill=H_FILL,align=CA)
        for i,p in enumerate(PERIODS,2): cs(ws.cell(row=row,column=i),phdr(p),font=H_FONT,fill=H_FILL,align=CA)
        cs(ws.cell(row=row,column=NCOLS),"← 新增 period",font=NEW_FONT,fill=NEW_FILL,align=CA)
        ws.row_dimensions[row].height=44

    def write_sec(ws,row,label):
        cs(ws.cell(row=row,column=1),f"  {label}",font=BD_FONT,fill=SEC_FILL,align=LA)
        for col in range(2,NCOLS+1):
            c=ws.cell(row=row,column=col); c.value=""; c.fill=SEC_FILL; c.border=BORDER

    def write_data(ws,row,label,mdict,is_total=False,num_fmt="#,##0",pct=False):
        fill=TOT_FILL if is_total else (ALT_FILL if row%2==0 else W_FILL)
        font=TOT_FONT if is_total else BD_FONT
        cs(ws.cell(row=row,column=1),label,font=font,fill=fill,align=LA)
        for i,p in enumerate(PERIODS,2):
            v=mdict.get(p); display=v if v is not None else "—"
            nf=("0.0%" if pct else num_fmt) if isinstance(v,(int,float)) else None
            cs(ws.cell(row=row,column=i),display,font=TOT_FONT if is_total else B_FONT,fill=fill,num_fmt=nf)
        cs(ws.cell(row=row,column=NCOLS),"",font=B_FONT,fill=NEW_FILL)

    def set_widths(ws,label_w,data_w):
        ws.column_dimensions["A"].width=label_w
        for i in range(2,NCOLS+1): ws.column_dimensions[get_column_letter(i)].width=data_w

    def title_row(ws,row,text,ncols=None):
        nc=ncols or NCOLS; ec=get_column_letter(nc)
        ws.merge_cells(f"A{row}:{ec}{row}")
        c=ws.cell(row=row,column=1,value=text)
        c.font=T_FONT; c.fill=SEC_FILL; c.alignment=CA; ws.row_dimensions[row].height=22

    def get_unit(key):
        if "pct" in key: return "ratio"
        if key.startswith("eps"): return "USD_per_share"
        if "shares" in key: return "millions_shares"
        return "USD_millions"

    seg = data["segment_data"]
    SEG_KEYS = ["CMBU","CDBU","MCBU","AEBU"]
    wb = Workbook()

    # Overview
    ws=wb.active; ws.title="Overview"; ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=58
    ws.merge_cells("A1:B1")
    c=ws["A1"]; c.value="Micron Technology (MU) — Financial Timeseries"
    c.font=T_FONT; c.fill=SEC_FILL; c.alignment=CA; ws.row_dimensions[1].height=28
    r=2
    for k,v in [("Company",data["metadata"]["company"]),
                ("Ticker",f"{data['metadata']['ticker']} / {data['metadata']['exchange']}"),
                ("Currency/Unit",f"{data['metadata']['currency']} / {data['metadata']['unit']}"),
                ("Last Updated",data["metadata"]["last_updated"]),
                ("Periods"," | ".join(PERIODS))]:
        for col,val in [(1,k),(2,v)]:
            c=ws.cell(row=r,column=col,value=val)
            c.font=BD_FONT if col==1 else B_FONT
            c.fill=ALT_FILL if r%2==0 else W_FILL; c.border=BORDER; c.alignment=LA
        r+=1
    r+=1
    ws.merge_cells(f"A{r}:B{r}")
    c=ws.cell(row=r,column=1,value="Filing Registry"); c.font=H_FONT; c.fill=H_FILL; c.alignment=CA; r+=1
    for p in PERIODS:
        f=data["filings"][p]
        for col,val in [(1,p),(2,f"{f['form']} | Period end: {f['period_end']} | Filed: {f['filing_date']}")]:
            c=ws.cell(row=r,column=col,value=val)
            c.font=BD_FONT if col==1 else B_FONT
            c.fill=ALT_FILL if r%2==0 else W_FILL; c.border=BORDER; c.alignment=LA
        r+=1
    r+=1; ws.merge_cells(f"A{r}:B{r}")
    c=ws.cell(row=r,column=1,value="▶ 如何新增季度 → 執行 update_financials.py")
    c.font=Font(name="Calibri",bold=True,color="375623",size=11); c.fill=NEW_FILL; c.alignment=LA; c.border=BORDER

    # Income Statement
    ws2=wb.create_sheet("Income Statement"); ws2.sheet_view.showGridLines=False; set_widths(ws2,38,16)
    r=1; title_row(ws2,r,"Consolidated Statements of Operations (USD Millions, except per share)"); r+=1
    write_col_hdrs(ws2,r); r+=1
    IS=data["income_statement"]
    for label,key,is_tot,nf,pct in [
        ("Revenue","revenue",True,"#,##0",False),
        ("Cost of Goods Sold","cost_of_goods_sold",False,"#,##0",False),
        ("Gross Margin","gross_margin",True,"#,##0",False),
        ("  Gross Margin %","gross_margin_pct",False,"0.0%",True),
        ("Research & Development","research_and_development",False,"#,##0",False),
        ("Selling, General & Administrative","selling_general_administrative",False,"#,##0",False),
        ("Other Operating (Income) Expense","other_operating_income_expense_net",False,"#,##0",False),
        ("Operating Income","operating_income",True,"#,##0",False),
        ("  Operating Margin %","operating_margin_pct",False,"0.0%",True),
        ("Interest Income (net)","interest_income_net",False,"#,##0",False),
        ("Other Non-op Income (Expense)","other_nonoperating_income_expense",False,"#,##0",False),
        ("Income Before Taxes","income_before_taxes",True,"#,##0",False),
        ("Income Tax Provision","income_tax_provision",False,"#,##0",False),
        ("  Effective Tax Rate","effective_tax_rate",False,"0.0%",True),
        ("Equity in Net Income of Investees","equity_in_net_income_of_investees",False,"#,##0",False),
        ("Net Income","net_income",True,"#,##0",False),
        ("  Net Margin %","net_margin_pct",False,"0.0%",True),
        ("EPS — Basic","eps_basic",False,"$#,##0.00",False),
        ("EPS — Diluted","eps_diluted",True,"$#,##0.00",False),
        ("Shares Basic (M)","shares_basic_millions",False,"#,##0",False),
        ("Shares Diluted (M)","shares_diluted_millions",False,"#,##0",False),
    ]:
        write_data(ws2,r,label,IS.get(key,{}),is_total=is_tot,num_fmt=nf,pct=pct); r+=1

    # Balance Sheet
    ws3=wb.create_sheet("Balance Sheet"); ws3.sheet_view.showGridLines=False; set_widths(ws3,40,16)
    r=1; title_row(ws3,r,"Consolidated Balance Sheets (USD Millions, as of period end)"); r+=1
    write_col_hdrs(ws3,r); r+=1
    TOTAL_BS={"total_current_assets","total_assets","total_current_liabilities",
              "total_liabilities","total_equity","total_liabilities_and_equity"}
    bs2=data["balance_sheet"]
    for sec_name,sec_key in [("ASSETS","assets"),("LIABILITIES","liabilities"),("EQUITY","equity")]:
        write_sec(ws3,r,sec_name); r+=1
        for key,vals in bs2[sec_key].items():
            write_data(ws3,r,key.replace("_"," ").title(),vals,is_total=(key in TOTAL_BS)); r+=1

    # Cash Flow
    ws4=wb.create_sheet("Cash Flow"); ws4.sheet_view.showGridLines=False; set_widths(ws4,44,16)
    r=1; title_row(ws4,r,"Consolidated Statements of Cash Flows (USD Millions, quarterly)"); r+=1
    write_col_hdrs(ws4,r); r+=1
    CF_TOTAL={"net_cash_from_operating_activities","net_cash_used_for_investing_activities",
              "net_cash_used_for_financing_activities"}
    cf2=data["cash_flow_statement"]
    for sec_name,sec_key in [("OPERATING ACTIVITIES","operating_activities"),
                              ("INVESTING ACTIVITIES","investing_activities"),
                              ("FINANCING ACTIVITIES","financing_activities")]:
        write_sec(ws4,r,sec_name); r+=1
        for key,vals in cf2[sec_key].items():
            write_data(ws4,r,key.replace("_"," ").title(),vals,is_total=(key in CF_TOTAL)); r+=1
    write_sec(ws4,r,"SUMMARY"); r+=1
    for key in ["fx_effect_on_cash","net_change_in_cash","beginning_cash","ending_cash","free_cash_flow"]:
        write_data(ws4,r,key.replace("_"," ").title(),cf2[key],
                   is_total=(key in {"net_change_in_cash","ending_cash","free_cash_flow"})); r+=1

    # Revenue Breakdown
    ws5=wb.create_sheet("Revenue Breakdown"); ws5.sheet_view.showGridLines=False; set_widths(ws5,32,16)
    r=1; title_row(ws5,r,"Revenue Breakdown — By Product & Business Unit (USD Millions)"); r+=1
    write_col_hdrs(ws5,r); r+=1
    rb2=data["revenue_by_product"]
    write_sec(ws5,r,"BY PRODUCT TECHNOLOGY"); r+=1
    for key in ["DRAM","NAND","Other_NOR","total","DRAM_mix_pct","NAND_mix_pct","Other_NOR_mix_pct"]:
        pct="pct" in key
        write_data(ws5,r,key.replace("_"," "),rb2.get(key,{}),is_total=(key=="total"),
                   num_fmt="0.0%" if pct else "#,##0",pct=pct); r+=1
    r+=1; write_sec(ws5,r,"BY BUSINESS UNIT"); r+=1
    for sk in SEG_KEYS:
        s=seg[sk]
        write_data(ws5,r,f"{sk} Revenue",s.get("revenue",{})); r+=1
        write_data(ws5,r,f"  {sk} Mix %",s.get("revenue_mix_pct",{}),num_fmt="0.0%",pct=True); r+=1
    write_data(ws5,r,"Total Revenue",seg["total"].get("revenue",{}),is_total=True); r+=1

    # Segment P&L
    ws6=wb.create_sheet("Segment P&L"); ws6.sheet_view.showGridLines=False; set_widths(ws6,36,16)
    r=1; title_row(ws6,r,"Segment P&L — Revenue & Operating Income by Business Unit (USD Millions)"); r+=1
    write_col_hdrs(ws6,r); r+=1
    for sk in SEG_KEYS:
        s=seg[sk]
        write_sec(ws6,r,f"{sk} — {s.get('full_name',sk)}"); r+=1
        write_data(ws6,r,"  Revenue",s.get("revenue",{})); r+=1
        write_data(ws6,r,"  Operating Income",s.get("operating_income",{})); r+=1
        write_data(ws6,r,"  Operating Margin %",s.get("operating_margin_pct",{}),num_fmt="0.0%",pct=True); r+=1
        write_data(ws6,r,"  D&A",s.get("depreciation_amortization",{})); r+=1
    write_sec(ws6,r,"CONSOLIDATED TOTAL"); r+=1
    write_data(ws6,r,"Total Revenue",         seg["total"].get("revenue",{}),          is_total=True); r+=1
    write_data(ws6,r,"Total Operating Income",seg["total"].get("operating_income",{}), is_total=True); r+=1
    write_data(ws6,r,"  Unallocated",          seg["total"].get("unallocated",{})); r+=1

    # Long Format
    ws7=wb.create_sheet("Long Format (DB)"); ws7.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFG",[14,14,12,28,42,16,14]): ws7.column_dimensions[col].width=w
    title_row(ws7,1,"Long Format — Tidy Data for Database / Looker Studio Import",ncols=7)
    for ci,h in enumerate(["period","period_end","form","category","metric","value","unit"],1):
        c=ws7.cell(row=2,column=ci,value=h); c.font=H_FONT; c.fill=H_FILL; c.alignment=CA; c.border=BORDER

    long_rows=[]
    for key,vals in data["income_statement"].items():
        for p in PERIODS:
            v=vals.get(p)
            if v is not None: long_rows.append([p,P_ENDS[p],P_FORMS[p],"income_statement",key,v,get_unit(key)])
    for sec_key,sec_dict in data["balance_sheet"].items():
        for key,vals in sec_dict.items():
            for p in PERIODS:
                v=vals.get(p)
                if v is not None: long_rows.append([p,P_ENDS[p],P_FORMS[p],f"balance_sheet_{sec_key}",key,v,"USD_millions"])
    for sec_key,sec_val in data["cash_flow_statement"].items():
        if isinstance(sec_val,dict) and all(isinstance(v,dict) for v in sec_val.values()):
            for key,vals in sec_val.items():
                for p in PERIODS:
                    v=vals.get(p)
                    if v is not None: long_rows.append([p,P_ENDS[p],P_FORMS[p],f"cash_flow_{sec_key}",key,v,"USD_millions"])
        else:
            for p in PERIODS:
                v=sec_val.get(p)
                if v is not None: long_rows.append([p,P_ENDS[p],P_FORMS[p],"cash_flow_summary",sec_key,v,"USD_millions"])
    for key,vals in data["revenue_by_product"].items():
        for p in PERIODS:
            v=vals.get(p)
            if v is not None: long_rows.append([p,P_ENDS[p],P_FORMS[p],"revenue_by_product",key,v,get_unit(key)])
    for sk in SEG_KEYS:
        s=seg[sk]
        for metric in ["revenue","operating_income","operating_margin_pct","depreciation_amortization"]:
            for p in PERIODS:
                v=s.get(metric,{}).get(p)
                if v is not None: long_rows.append([p,P_ENDS[p],P_FORMS[p],f"segment_{sk}",metric,v,get_unit(metric)])

    for i,row in enumerate(long_rows,3):
        fill=ALT_FILL if i%2==0 else W_FILL
        for ci,val in enumerate(row,1):
            c=ws7.cell(row=i,column=ci,value=val)
            c.font=B_FONT; c.fill=fill; c.border=BORDER
            c.alignment=RA if ci==6 else LA
            if ci==6 and isinstance(val,(int,float)): c.number_format="#,##0.####"

    # Notable Events
    ws8=wb.create_sheet("Notable Events"); ws8.sheet_view.showGridLines=False
    for col,w in zip("ABCD",[14,22,14,80]): ws8.column_dimensions[col].width=w
    title_row(ws8,1,"Notable Events (可持續追加列)",ncols=4)
    for ci,h in enumerate(["period","category","date","description"],1):
        c=ws8.cell(row=2,column=ci,value=h); c.font=H_FONT; c.fill=H_FILL; c.alignment=CA; c.border=BORDER
    for i,ev in enumerate(data["notable_events"],3):
        fill=ALT_FILL if i%2==0 else W_FILL
        for ci,val in enumerate([ev["period"],ev["category"],ev["date"],ev["description"]],1):
            c=ws8.cell(row=i,column=ci,value=val)
            c.font=B_FONT; c.fill=fill; c.border=BORDER
            c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=(ci==4))
        ws8.row_dimensions[i].height=36

    wb.save(xlsx_path)
    return len(long_rows)


# ══════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"載入 {JSON_PATH} ...")
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    existing = data["metadata"]["periods"]
    if NEW_PERIOD in existing:
        print(f"⚠️  {NEW_PERIOD} 已存在，將覆蓋現有數據。")
    else:
        print(f"✓ 新增期間：{NEW_PERIOD}")

    # 備份
    backup_path = JSON_PATH.replace(".json", f"_backup_{date.today()}.json")
    shutil.copy(JSON_PATH, backup_path)
    print(f"✓ 備份 JSON → {backup_path}")

    # 合併
    data = merge_period(
        data,
        period    = NEW_PERIOD,
        filing    = NEW_FILING,
        is_dict   = NEW_INCOME_STATEMENT,
        bs        = NEW_BALANCE_SHEET,
        cf        = NEW_CASH_FLOW,
        rb        = NEW_REVENUE_BY_PRODUCT,
        seg_new   = NEW_SEGMENT_DATA,
        events    = NEW_NOTABLE_EVENTS,
    )

    # 儲存 JSON
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"✓ JSON 已更新：{JSON_PATH}")

    # 重建 XLSX
    n = rebuild_xlsx(data, XLSX_PATH)
    print(f"✓ XLSX 已重建：{XLSX_PATH}  (Long Format: {n} 列)")
    print(f"\n完成！現有期間：{' | '.join(data['metadata']['periods'])}")
