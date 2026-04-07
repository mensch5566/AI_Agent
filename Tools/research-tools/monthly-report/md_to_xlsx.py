"""
Stock Monthly: Obsidian .md → Excel (.xlsx)

用法:
  uv run --with openpyxl python3 md_to_xlsx.py [--month 2026-03] [--out Stock_Monthly.xlsx]

預設讀取所有標的的最新月報，輸出到 Stock_Monthly 資料夾。
"""

import argparse
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 路徑 ──────────────────────────────────────────────
OBSIDIAN_BASE = Path.home() / "Library/Mobile Documents/iCloud~md~obsidian/Documents/Obsidian/Khouse/Semiconductors"
OUTPUT_DIR = Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Assembla/trunk/喬大地產/Stock_Monthly"

# 標的 → Obsidian 資料夾名
TICKERS = {
    "SNDK": "SNDK",
    "NVDA": "NVDA",
    "MU": "MU",
    "聯發科": "聯發科",
}

# ── 樣式 ──────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ── 解析 Markdown ────────────────────────────────────
def parse_frontmatter(text: str) -> dict:
    """解析 YAML frontmatter。"""
    m = re.match(r"^---\s*\n(.*?)\n---", text, re.DOTALL)
    if not m:
        return {}
    meta = {}
    for line in m.group(1).strip().splitlines():
        if ":" in line:
            key, val = line.split(":", 1)
            val = val.strip()
            # 嘗試轉數字
            if val:
                try:
                    val = float(val.replace("%", ""))
                    if val == int(val):
                        val = int(val)
                except ValueError:
                    pass
            else:
                val = None
            meta[key.strip()] = val
    return meta


def parse_table(section_text: str) -> list[list[str]]:
    """解析 Markdown 表格，回傳 [header, row1, row2, ...]。"""
    rows = []
    for line in section_text.strip().splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        # 跳過分隔線 |---|---|
        if re.match(r"^\|[\s\-:|]+\|$", line):
            continue
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if any(c for c in cells):  # 至少一個非空
            rows.append(cells)
    return rows


def parse_sections(text: str) -> dict[str, str]:
    """把 Markdown 拆成 {section_title: section_body}。"""
    sections = {}
    parts = re.split(r"^## (.+)$", text, flags=re.MULTILINE)
    # parts[0] = frontmatter 之前的內容, parts[1] = title1, parts[2] = body1, ...
    for i in range(1, len(parts), 2):
        title = parts[i].strip()
        body = parts[i + 1] if i + 1 < len(parts) else ""
        sections[title] = body
    return sections


def parse_monthly_md(filepath: Path) -> dict:
    """解析一份月報 .md，回傳結構化資料。"""
    text = filepath.read_text(encoding="utf-8")
    meta = parse_frontmatter(text)
    sections = parse_sections(text)

    result = {"meta": meta, "kpi": [], "risk": [], "base_risk": []}

    for title, body in sections.items():
        table = parse_table(body)
        if not table:
            continue
        # 跳過 header row
        data_rows = table[1:] if len(table) > 1 else []
        if "KPI" in title:
            result["kpi"] = data_rows
        elif "目標價風險" in title:
            result["risk"] = data_rows
        elif "底層風險" in title:
            result["base_risk"] = data_rows

    return result


# ── 寫 Excel ─────────────────────────────────────────
def write_header_row(ws, row: int, headers: list[str], col_start: int = 1):
    """寫表頭列，套用樣式。"""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def write_data_rows(ws, start_row: int, rows: list[list[str]], col_start: int = 1):
    """寫資料列，套用邊框。"""
    for r_idx, row_data in enumerate(rows):
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=start_row + r_idx, column=col_start + c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_sheet(ws, data: dict):
    """把解析好的資料寫入一個 sheet。"""
    meta = data["meta"]
    row = 1

    # ── Frontmatter 區域 ──
    info_items = [
        ("標的", meta.get("ticker", "")),
        ("成本", meta.get("成本")),
        ("目標價", meta.get("目標價")),
        ("預期漲幅", meta.get("預期漲幅")),
    ]

    write_header_row(ws, row, ["指標", "值"])
    row += 1
    for label, val in info_items:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.border = THIN_BORDER
        # 成本和目標價格式化
        if label in ("成本", "目標價") and isinstance(val, (int, float)):
            cell_val.number_format = '"$"#,##0.00'
        elif label == "預期漲幅" and isinstance(val, (int, float)):
            cell_val.value = val / 100 if val > 1 else val
            cell_val.number_format = "0.0%"
        row += 1

    row += 1  # 空行

    # ── 各 section ──
    sections = [
        ("目標價 KPI", ["KPI", "描述", "狀態"], data["kpi"]),
        ("目標價風險", ["風險", "描述", "狀態"], data["risk"]),
        ("底層風險", ["風險", "描述", "狀態"], data["base_risk"]),
    ]

    for title, headers, rows_data in sections:
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        row += 1
        write_header_row(ws, row, headers)
        row += 1
        write_data_rows(ws, row, rows_data)
        row += len(rows_data)
        row += 1  # 空行

    # ── 欄寬 ──
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12


def find_latest_monthly(ticker_dir: Path, month: str | None) -> Path | None:
    """找到指定月份或最新的月報 .md。"""
    monthly_dir = ticker_dir / "Stock_Monthly"
    if not monthly_dir.exists():
        return None

    if month:
        target = monthly_dir / f"{month}.md"
        return target if target.exists() else None

    # 找最新的
    files = sorted(monthly_dir.glob("*.md"), reverse=True)
    return files[0] if files else None


def main():
    parser = argparse.ArgumentParser(description="Stock Monthly .md → .xlsx")
    parser.add_argument("--month", help="指定月份，如 2026-03")
    parser.add_argument("--out", help="輸出檔名", default=None)
    args = parser.parse_args()

    month_label = args.month or "latest"
    out_name = args.out or f"Stock_Monthly_{args.month}.xlsx" if args.month else "Stock_Monthly.xlsx"
    out_path = OUTPUT_DIR / out_name

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除預設 sheet

    found = 0
    for sheet_name, folder_name in TICKERS.items():
        ticker_dir = OBSIDIAN_BASE / folder_name
        md_path = find_latest_monthly(ticker_dir, args.month)
        if not md_path:
            print(f"⚠ {sheet_name}: 找不到 {month_label} 月報，跳過")
            continue

        print(f"✓ {sheet_name}: {md_path.name}")
        data = parse_monthly_md(md_path)
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, data)
        found += 1

    if found == 0:
        print("沒有找到任何月報 .md，未產生 Excel。")
        return

    wb.save(out_path)
    print(f"\n✅ 已輸出: {out_path}")


if __name__ == "__main__":
    main()
