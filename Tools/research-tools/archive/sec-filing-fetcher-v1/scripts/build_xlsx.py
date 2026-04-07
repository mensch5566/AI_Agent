"""
build_xlsx.py — 由 timeseries JSON 重建 XLSX（8 個 Sheet）
用法：
    python3 build_xlsx.py \
      --json /path/to/MU_financials.json \
      --out  /path/to/MU_financials.xlsx
"""

import argparse, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 樣式常數 ──────────────────────────────────────────────────
H_FILL  = PatternFill("solid", fgColor="1F4E79")
SEC_FILL= PatternFill("solid", fgColor="D6E4F0")
TOT_FILL= PatternFill("solid", fgColor="FFF2CC")
ALT_FILL= PatternFill("solid", fgColor="F2F8FD")
W_FILL  = PatternFill("solid", fgColor="FFFFFF")
NEW_FILL= PatternFill("solid", fgColor="E2EFDA")

H_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
T_FONT  = Font(name="Calibri", bold=True, color="1F4E79", size=13)
B_FONT  = Font(name="Calibri", size=10)
BD_FONT = Font(name="Calibri", bold=True, size=10)
TOT_FONT= Font(name="Calibri", bold=True, color="7B3F00", size=10)
NEW_FONT= Font(name="Calibri", bold=True, color="375623", size=9)

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CA = Alignment(horizontal="center", vertical="center", wrap_text=True)
RA = Alignment(horizontal="right",  vertical="center")
LA = Alignment(horizontal="left",   vertical="center")


class XLSXBuilder:
    def __init__(self, data):
        self.data    = data
        self.periods = data["metadata"]["periods"]
        self.p_ends  = {p: data["filings"][p]["period_end"] for p in self.periods}
        self.p_forms = {p: data["filings"][p]["form"]       for p in self.periods}
        self.ncols   = len(self.periods) + 2  # label + periods + 預留

    def cs(self, c, value, font=None, fill=None, align=None, num_fmt=None):
        font = font or B_FONT; fill = fill or W_FILL; align = align or RA
        c.value = value; c.font = font; c.fill = fill
        c.alignment = align; c.border = BORDER
        if num_fmt and isinstance(value, (int, float)): c.number_format = num_fmt

    def phdr(self, p):
        return f"{p}\n{self.p_ends[p]}\n({self.p_forms[p]})"

    def col_headers(self, ws, row):
        self.cs(ws.cell(row=row, column=1), "", font=H_FONT, fill=H_FILL, align=CA)
        for i, p in enumerate(self.periods, 2):
            self.cs(ws.cell(row=row, column=i), self.phdr(p), font=H_FONT, fill=H_FILL, align=CA)
        self.cs(ws.cell(row=row, column=self.ncols), "← 新增 period",
                font=NEW_FONT, fill=NEW_FILL, align=CA)
        ws.row_dimensions[row].height = 44

    def section(self, ws, row, label):
        self.cs(ws.cell(row=row, column=1), f"  {label}", font=BD_FONT, fill=SEC_FILL, align=LA)
        for col in range(2, self.ncols + 1):
            c = ws.cell(row=row, column=col); c.value = ""; c.fill = SEC_FILL; c.border = BORDER

    def data_row(self, ws, row, label, mdict, is_total=False, num_fmt="#,##0", pct=False):
        fill = TOT_FILL if is_total else (ALT_FILL if row % 2 == 0 else W_FILL)
        font = TOT_FONT if is_total else BD_FONT
        self.cs(ws.cell(row=row, column=1), label, font=font, fill=fill, align=LA)
        for i, p in enumerate(self.periods, 2):
            v = (mdict or {}).get(p)
            nf = ("0.0%" if pct else num_fmt) if isinstance(v, (int, float)) else None
            self.cs(ws.cell(row=row, column=i), v if v is not None else "—",
                    font=TOT_FONT if is_total else B_FONT, fill=fill, num_fmt=nf)
        self.cs(ws.cell(row=row, column=self.ncols), "", font=B_FONT, fill=NEW_FILL)

    def set_widths(self, ws, label_w, data_w):
        ws.column_dimensions["A"].width = label_w
        for i in range(2, self.ncols + 1):
            ws.column_dimensions[get_column_letter(i)].width = data_w

    def title(self, ws, row, text, ncols=None):
        nc = ncols or self.ncols
        ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = T_FONT; c.fill = SEC_FILL; c.alignment = CA
        ws.row_dimensions[row].height = 22

    def get_unit(self, key):
        if "pct" in key: return "ratio"
        if key.startswith("eps"): return "USD_per_share"
        if "shares" in key: return "millions_shares"
        return "USD_millions"

    # ── 各 Sheet ────────────────────────────────────────────

    def sheet_overview(self, wb):
        ws = wb.active; ws.title = "Overview"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 58
        ws.merge_cells("A1:B1")
        c = ws["A1"]
        c.value = f"{self.data['metadata']['ticker']} — Financial Timeseries"
        c.font = T_FONT; c.fill = SEC_FILL; c.alignment = CA
        ws.row_dimensions[1].height = 28
        r = 2
        m = self.data["metadata"]
        for k, v in [
            ("Company",       m["company"]),
            ("Ticker",        f"{m['ticker']} / {m['exchange']}"),
            ("Currency/Unit", f"{m['currency']} / {m['unit']}"),
            ("Last Updated",  m["last_updated"]),
            ("Periods",       " | ".join(self.periods)),
        ]:
            for col, val in [(1, k), (2, v)]:
                c = ws.cell(row=r, column=col, value=val)
                c.font = BD_FONT if col == 1 else B_FONT
                c.fill = ALT_FILL if r % 2 == 0 else W_FILL
                c.border = BORDER; c.alignment = LA
            r += 1
        r += 1
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value="Filing Registry")
        c.font = H_FONT; c.fill = H_FILL; c.alignment = CA; r += 1
        for p in self.periods:
            f = self.data["filings"][p]
            for col, val in [(1, p), (2, f"{f['form']} | {f['period_end']} | Filed: {f['filing_date']}")]:
                c = ws.cell(row=r, column=col, value=val)
                c.font = BD_FONT if col == 1 else B_FONT
                c.fill = ALT_FILL if r % 2 == 0 else W_FILL
                c.border = BORDER; c.alignment = LA
            r += 1

    def sheet_income(self, wb):
        ws = wb.create_sheet("Income Statement")
        ws.sheet_view.showGridLines = False
        self.set_widths(ws, 38, 16)
        r = 1; self.title(ws, r, "Consolidated Statements of Operations (USD Millions, except per share)"); r += 1
        self.col_headers(ws, r); r += 1
        IS = self.data.get("income_statement", {})
        for label, key, is_tot, nf, pct in [
            ("Revenue",                           "revenue",                            True,  "#,##0", False),
            ("Cost of Goods Sold",                "cost_of_goods_sold",                 False, "#,##0", False),
            ("Gross Margin",                      "gross_margin",                       True,  "#,##0", False),
            ("  Gross Margin %",                  "gross_margin_pct",                   False, "0.0%",  True),
            ("Research & Development",            "research_and_development",           False, "#,##0", False),
            ("Selling, General & Administrative", "selling_general_administrative",     False, "#,##0", False),
            ("Other Operating (Income) Expense",  "other_operating_income_expense_net", False, "#,##0", False),
            ("Operating Income",                  "operating_income",                   True,  "#,##0", False),
            ("  Operating Margin %",              "operating_margin_pct",               False, "0.0%",  True),
            ("Interest Income (net)",             "interest_income_net",                False, "#,##0", False),
            ("Other Non-op Income (Expense)",     "other_nonoperating_income_expense",  False, "#,##0", False),
            ("Income Before Taxes",               "income_before_taxes",                True,  "#,##0", False),
            ("Income Tax Provision",              "income_tax_provision",               False, "#,##0", False),
            ("  Effective Tax Rate",              "effective_tax_rate",                 False, "0.0%",  True),
            ("Equity in Net Income of Investees", "equity_in_net_income_of_investees",  False, "#,##0", False),
            ("Net Income",                        "net_income",                         True,  "#,##0", False),
            ("  Net Margin %",                    "net_margin_pct",                     False, "0.0%",  True),
            ("EPS — Basic",                       "eps_basic",                          False, "$#,##0.00", False),
            ("EPS — Diluted",                     "eps_diluted",                        True,  "$#,##0.00", False),
            ("Shares Basic (M)",                  "shares_basic_millions",              False, "#,##0", False),
            ("Shares Diluted (M)",                "shares_diluted_millions",            False, "#,##0", False),
        ]:
            self.data_row(ws, r, label, IS.get(key), is_total=is_tot, num_fmt=nf, pct=pct); r += 1

    def sheet_balance(self, wb):
        ws = wb.create_sheet("Balance Sheet")
        ws.sheet_view.showGridLines = False
        self.set_widths(ws, 40, 16)
        r = 1; self.title(ws, r, "Consolidated Balance Sheets (USD Millions, as of period end)"); r += 1
        self.col_headers(ws, r); r += 1
        TOTAL_KEYS = {"total_current_assets","total_assets","total_current_liabilities",
                      "total_liabilities","total_equity","total_liabilities_and_equity"}
        bs = self.data.get("balance_sheet", {})
        for sec_name, sec_key in [("ASSETS","assets"),("LIABILITIES","liabilities"),("EQUITY","equity")]:
            self.section(ws, r, sec_name); r += 1
            for key, vals in bs.get(sec_key, {}).items():
                self.data_row(ws, r, key.replace("_"," ").title(), vals, is_total=(key in TOTAL_KEYS)); r += 1

    def sheet_cashflow(self, wb):
        ws = wb.create_sheet("Cash Flow")
        ws.sheet_view.showGridLines = False
        self.set_widths(ws, 44, 16)
        r = 1; self.title(ws, r, "Consolidated Statements of Cash Flows (USD Millions, quarterly)"); r += 1
        self.col_headers(ws, r); r += 1
        CF_TOT = {"net_cash_from_operating_activities","net_cash_used_for_investing_activities",
                  "net_cash_used_for_financing_activities"}
        cf = self.data.get("cash_flow_statement", {})
        for sec_name, sec_key in [("OPERATING ACTIVITIES","operating_activities"),
                                   ("INVESTING ACTIVITIES","investing_activities"),
                                   ("FINANCING ACTIVITIES","financing_activities")]:
            self.section(ws, r, sec_name); r += 1
            for key, vals in cf.get(sec_key, {}).items():
                self.data_row(ws, r, key.replace("_"," ").title(), vals, is_total=(key in CF_TOT)); r += 1
        self.section(ws, r, "SUMMARY"); r += 1
        for key in ["fx_effect_on_cash","net_change_in_cash","beginning_cash","ending_cash","free_cash_flow"]:
            is_tot = key in {"net_change_in_cash","ending_cash","free_cash_flow"}
            self.data_row(ws, r, key.replace("_"," ").title(), cf.get(key,{}), is_total=is_tot); r += 1

    def sheet_revenue(self, wb):
        ws = wb.create_sheet("Revenue Breakdown")
        ws.sheet_view.showGridLines = False
        self.set_widths(ws, 32, 16)
        r = 1; self.title(ws, r, "Revenue Breakdown — By Product & Business Unit (USD Millions)"); r += 1
        self.col_headers(ws, r); r += 1
        rb = self.data.get("revenue_by_product", {})
        self.section(ws, r, "BY PRODUCT TECHNOLOGY"); r += 1
        for key, vals in rb.items():
            pct = "pct" in key; is_tot = key == "total"
            self.data_row(ws, r, key.replace("_"," "), vals,
                          is_total=is_tot, num_fmt="0.0%" if pct else "#,##0", pct=pct); r += 1
        r += 1
        seg = self.data.get("segment_data", {})
        self.section(ws, r, "BY BUSINESS UNIT"); r += 1
        for sk in [k for k in seg if k != "total"]:
            s = seg[sk]
            self.data_row(ws, r, f"{sk} Revenue", s.get("revenue", {})); r += 1
            self.data_row(ws, r, f"  {sk} Mix %", s.get("revenue_mix_pct", {}), num_fmt="0.0%", pct=True); r += 1
        self.data_row(ws, r, "Total Revenue", seg.get("total",{}).get("revenue",{}), is_total=True); r += 1

    def sheet_segment(self, wb):
        ws = wb.create_sheet("Segment P&L")
        ws.sheet_view.showGridLines = False
        self.set_widths(ws, 36, 16)
        r = 1; self.title(ws, r, "Segment P&L — Revenue & Operating Income by Business Unit (USD Millions)"); r += 1
        self.col_headers(ws, r); r += 1
        seg = self.data.get("segment_data", {})
        for sk in [k for k in seg if k != "total"]:
            s = seg[sk]
            self.section(ws, r, f"{sk} — {s.get('full_name', sk)}"); r += 1
            self.data_row(ws, r, "  Revenue",            s.get("revenue",{})); r += 1
            self.data_row(ws, r, "  Operating Income",   s.get("operating_income",{})); r += 1
            self.data_row(ws, r, "  Operating Margin %", s.get("operating_margin_pct",{}), num_fmt="0.0%", pct=True); r += 1
            self.data_row(ws, r, "  D&A",                s.get("depreciation_amortization",{})); r += 1
        tot = seg.get("total", {})
        self.section(ws, r, "CONSOLIDATED TOTAL"); r += 1
        self.data_row(ws, r, "Total Revenue",          tot.get("revenue",{}),          is_total=True); r += 1
        self.data_row(ws, r, "Total Operating Income", tot.get("operating_income",{}), is_total=True); r += 1
        self.data_row(ws, r, "  Unallocated",           tot.get("unallocated",{})); r += 1

    def sheet_long_format(self, wb):
        ws = wb.create_sheet("Long Format (DB)")
        ws.sheet_view.showGridLines = False
        for col, w in zip("ABCDEFG", [14,14,12,28,42,16,14]):
            ws.column_dimensions[col].width = w
        self.title(ws, 1, "Long Format — Tidy Data for Database / Looker Studio Import", ncols=7)
        for ci, h in enumerate(["period","period_end","form","category","metric","value","unit"], 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = H_FONT; c.fill = H_FILL; c.alignment = CA; c.border = BORDER

        rows = []
        def add(p, cat, metric, val):
            if val is not None:
                rows.append([p, self.p_ends[p], self.p_forms[p], cat, metric, val, self.get_unit(metric)])

        for key, vals in self.data.get("income_statement", {}).items():
            for p in self.periods: add(p, "income_statement", key, vals.get(p))
        for sec, sec_d in self.data.get("balance_sheet", {}).items():
            for key, vals in sec_d.items():
                for p in self.periods: add(p, f"balance_sheet_{sec}", key, vals.get(p))
        for sec, sec_v in self.data.get("cash_flow_statement", {}).items():
            if isinstance(sec_v, dict) and all(isinstance(v, dict) for v in sec_v.values()):
                for key, vals in sec_v.items():
                    for p in self.periods: add(p, f"cash_flow_{sec}", key, vals.get(p))
            else:
                for p in self.periods: add(p, "cash_flow_summary", sec, (sec_v or {}).get(p))
        for key, vals in self.data.get("revenue_by_product", {}).items():
            for p in self.periods: add(p, "revenue_by_product", key, vals.get(p))
        for sk, s in self.data.get("segment_data", {}).items():
            if sk == "total": continue
            for metric in ["revenue","operating_income","operating_margin_pct","depreciation_amortization"]:
                for p in self.periods: add(p, f"segment_{sk}", metric, (s.get(metric) or {}).get(p))

        for i, row in enumerate(rows, 3):
            fill = ALT_FILL if i % 2 == 0 else W_FILL
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=i, column=ci, value=val)
                c.font = B_FONT; c.fill = fill; c.border = BORDER
                c.alignment = RA if ci == 6 else LA
                if ci == 6 and isinstance(val, (int, float)): c.number_format = "#,##0.####"
        return len(rows)

    def sheet_events(self, wb):
        ws = wb.create_sheet("Notable Events")
        ws.sheet_view.showGridLines = False
        for col, w in zip("ABCD", [14,22,14,80]): ws.column_dimensions[col].width = w
        self.title(ws, 1, "Notable Events (可持續追加列)", ncols=4)
        for ci, h in enumerate(["period","category","date","description"], 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = H_FONT; c.fill = H_FILL; c.alignment = CA; c.border = BORDER
        for i, ev in enumerate(self.data.get("notable_events", []), 3):
            fill = ALT_FILL if i % 2 == 0 else W_FILL
            for ci, val in enumerate([ev["period"],ev["category"],ev["date"],ev["description"]], 1):
                c = ws.cell(row=i, column=ci, value=val)
                c.font = B_FONT; c.fill = fill; c.border = BORDER
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(ci==4))
            ws.row_dimensions[i].height = 36

    def build(self, out_path):
        wb = Workbook()
        self.sheet_overview(wb)
        self.sheet_income(wb)
        self.sheet_balance(wb)
        self.sheet_cashflow(wb)
        self.sheet_revenue(wb)
        self.sheet_segment(wb)
        n_rows = self.sheet_long_format(wb)
        self.sheet_events(wb)
        wb.save(out_path)
        return n_rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", required=True, help="JSON 輸入路徑")
    parser.add_argument("--out",  required=True, help="XLSX 輸出路徑")
    args = parser.parse_args()

    with open(args.json) as f:
        data = json.load(f)

    builder = XLSXBuilder(data)
    n = builder.build(args.out)
    print(f"✓ XLSX saved: {args.out}")
    print(f"  Sheets: 8 | Long Format rows: {n}")
    print(f"  Periods: {' | '.join(data['metadata']['periods'])}")


if __name__ == "__main__":
    main()
